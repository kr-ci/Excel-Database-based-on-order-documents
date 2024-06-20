#pdfToExcelDataBase.py - programm parse content of each pdf files in current
#directory, lookin for specyfing line and put them in to excel database

import PyPDF2, openpyxl, os, re
from char_replacer import polishchar
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment
#Create xls file with specyfing headlines

wb=openpyxl.Workbook()
sheet=wb.active
heads=['Nr pliku','Data zlecenia','NIP','Nazwa zleceniobiorcy','Kierunek z','Kierunek do','Ref załadunku',
       'Miejscowość załadunku','Miejscowość rozładunku','Cena USD',
       'Dystans km','Cena/km']

for i in range(0,len(heads)):
    sheet[(get_column_letter(i+1))+'1']=heads[(i)]
    sheet.column_dimensions[(get_column_letter(i+1))].width = 25
    bo=Font(bold=True)
    sheet[(get_column_letter(i+1))+'1'].font=bo
    sheet[(get_column_letter(i+1))+'1'].alignment=Alignment(horizontal='center')

#searching for pdf files in current catalog
    
pdfFiles=[]
for filename in os.listdir('.'):
    if filename.endswith('.pdf'):
        pdfFiles.append(filename)
print('Stworzono listę plików PDF do sczytania.')

#open pdf files
num=2
for i in range(0, len(pdfFiles)):
    pdfFileObj=open(pdfFiles[i], 'rb')
    pdfReader=PyPDF2.PdfReader(pdfFileObj)
    pageObj=pdfReader.pages[0]
    x=pageObj.extract_text()
    pageObj1=pdfReader.pages[1]
    w=pageObj1.extract_text()
    pdfFileObj.close()
    x=x+w
    a=pdfFiles[i]
    a=a[:-4]
#lookin' for data

    #Data zlecenia

    try:
        dataRegex=re.compile(r'(?<=\nParzymiechy dolne, )(\d)*(\.)(\d)*(\.)(\d)*(?=\nZlecenie)',re.IGNORECASE)
        data=dataRegex.search(x)
        b=data.group()
    except:
        b='Puste'

    #NIP zleceniobiorcy

    try:
        nipRegex=re.compile(r'(?<=\nNIP : )(PL)*(\s)*(\d)*(?=\n)',re.IGNORECASE)
        nip=nipRegex.search(x)
        c=nip.group()
        dl=len(c)
        if dl>10:
            diff=dl-10
            c=c[diff:]
    except:
        c='Puste'

    #Nazwa zleceniobiorcy
    
    try:
        zlecRegex=re.compile(r'(?<=Zleceniobiorca)[\S\s]*(?=NIP|Towar|Trasa)')
        zlec=zlecRegex.search(x)
        d=zlec.group()
    except:
        d='Puste'

    #kierunek z
    try:
        kierzRegex=re.compile(r'(?<=((Załadunek)\n/))\w\w(?=/)')
        kierz=kierzRegex.search(x)
        e=kierz.group()
    except:
        e='Puste'

    #kierunek do

    try:
        kierdRegex=re.compile(r'(?<=((Rozładunek)\n/))\w\w(?=/)')
        kierd=kierdRegex.search(x)
        f=kierd.group()
    except:
        f='Puste'
    
    #Nr ref załadunku

    try:
        kodRegex=re.compile(r'(?<=(ref)[(\D){,3}])(.*)(?=\n)',re.IGNORECASE)
        kod=kodRegex.search(x)
        g=kod.group()
        a1=kod.group()
        while True:
            if a1.startswith(':'):
                a1=a1[1:]
            elif a1.startswith(' '):
                a1=a1[1:]
            elif a1.startswith('-'):
                a1=a1[1:]
            else:
                g=a1
                break
    except:
        g='Puste'


    #Miejsce załadunku

    try:
        miejzaRegex=re.compile(r'(?<=Załadunek\n/'+re.escape(e)+'(/))[\S\s]*(?=2Rozładunek\n)',re.IGNORECASE)
        miejza=miejzaRegex.search(x)
        h=miejza.group()
    except:
        h='Puste'

    #Mijesce rozładunku

    try:
        miejroRegex=re.compile(r'(?<=2Rozładunek\n/'+re.escape(f)+'(/))[\S\s]*(?=Waluta frachtu)',re.IGNORECASE)
        miejro=miejroRegex.search(x)
        i=miejro.group()    
    except:
        i='Puste'

    #Cena

    try:
        cenaRegex=re.compile(r'(?<=fracht 1,00 )(.*)(?=( USD ))')
        cena=cenaRegex.search(x)
        j=cena.group()
    except:
        j='Puste'
        
# put on to the list
    k=0
    li=[]
    li=a,b,c,d,e,f,g,h,i,j,k
# put list to excel
    for y in range(0,len(li)):
        sheet[(get_column_letter(y+1))+str(num)]=li[(y)]
        sheet[(get_column_letter(y+1))+str(num)].alignment=Alignment(horizontal='center',wrapText=True,vertical='center')
    num+=1
for v in range(2,(sheet.max_row+1)):
    v=str(v)
    sheet['L'+v].value='=J'+v+'/K'+v
    sheet['L'+v].alignment=Alignment(horizontal='center',wrapText=True,vertical='center')
print('Trwa eksport do bazy...')
wb.save('database.xlsx')
wb.close()
print('Gotowe!')
